from flask import Blueprint, request, jsonify, send_file, make_response
from io import BytesIO
from datetime import datetime
from api.models import db, Purchase, Event
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/reportes/ventas', methods=['GET'])
def ventas_report():
    """Endpoint que devuelve reportes de ventas en JSON o como archivo (PDF/Excel)"""
    try:
        # Leer filtros
        evento_id = request.args.get('evento_id')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        sector_id = request.args.get('sector_id')
        formato = request.args.get('formato')  # 'json' | 'pdf' | 'excel'

        # Construir consulta base: compras con estado diferente a 'cancelled' (por ejemplo)
        query = Purchase.query

        if evento_id:
            query = query.filter(Purchase.event_id == str(evento_id))

        if fecha_inicio:
            try:
                dt_inicio = datetime.fromisoformat(fecha_inicio)
                query = query.filter(Purchase.purchase_date >= dt_inicio)
            except Exception:
                pass
        if fecha_fin:
            try:
                dt_fin = datetime.fromisoformat(fecha_fin)
                query = query.filter(Purchase.purchase_date <= dt_fin)
            except Exception:
                pass

        purchases = query.order_by(Purchase.purchase_date.desc()).all()

        # Agregar cálculos para el reporte
        total_ventas = sum(p.total_price for p in purchases)
        total_entradas = sum(p.quantity for p in purchases)
        promedio_venta = (total_ventas / len(purchases)) if purchases else 0

        # Análisis por sector (usamos Event.category)
        analisis_por_sector = {}
        analisis_por_evento = {}
        datos_detallados = []

        for p in purchases:
            evento = p.event
            sector = (evento.category or 'General') if evento else 'General'
            if sector not in analisis_por_sector:
                analisis_por_sector[sector] = {
                    'entradas_vendidas': 0,
                    'total_ventas': 0.0,
                    'precio_promedio': 0.0
                }

            analisis_por_sector[sector]['entradas_vendidas'] += p.quantity
            analisis_por_sector[sector]['total_ventas'] += p.total_price

            # Evento
            evento_nombre = evento.title if evento else f'Evento {p.event_id}'
            if evento_nombre not in analisis_por_evento:
                analisis_por_evento[evento_nombre] = {
                    'total_ventas': 0.0,
                    'total_entradas': 0
                }
            analisis_por_evento[evento_nombre]['total_ventas'] += p.total_price
            analisis_por_evento[evento_nombre]['total_entradas'] += p.quantity

            datos_detallados.append({
                'id': p.id,
                'fecha_venta': p.purchase_date.isoformat() if p.purchase_date else None,
                'cantidad': p.quantity,
                'precio_unitario': p.unit_price,
                'total': p.total_price,
                'cliente_nombre': p.user.name if p.user else 'Cliente',
                'cliente_rut': getattr(p.user, 'rut', ''),
                'metodo_pago': 'online',
                'evento_nombre': evento.title if evento else '',
                'fecha_evento': evento.date if evento else '',
                'lugar': evento.venue if evento else '',
                'sector_nombre': sector,
            })

        # Calcular precio_promedio por sector
        for sector, data in analisis_por_sector.items():
            entradas = data['entradas_vendidas']
            data['precio_promedio'] = (data['total_ventas'] / entradas) if entradas else 0

        resumen_ejecutivo = {
            'total_ventas': total_ventas,
            'total_entradas': total_entradas,
            'promedio_venta': promedio_venta,
            'sector_mas_vendido': max(analisis_por_sector.items(), key=lambda x: x[1]['entradas_vendidas'])[0] if analisis_por_sector else None,
            'sector_mayor_ingreso': max(analisis_por_sector.items(), key=lambda x: x[1]['total_ventas'])[0] if analisis_por_sector else None
        }

        response_json = {
            'success': True,
            'resumen_ejecutivo': resumen_ejecutivo,
            'analisis_por_sector': analisis_por_sector,
            'analisis_por_evento': analisis_por_evento,
            'datos_detallados': datos_detallados,
            'filtros_aplicados': {
                'evento_id': evento_id,
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'sector_id': sector_id
            },
            'total_registros': len(datos_detallados)
        }

        # Si piden un formato de archivo, devolvemos PDF o Excel real
        if formato == 'pdf':
            return generar_pdf_reporte(response_json)
        elif formato == 'excel':
            return generar_excel_reporte(response_json)

        return jsonify(response_json)

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def generar_pdf_reporte(data):
    """Genera un PDF profesional con los datos del reporte"""
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        # Container para los elementos del PDF
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Título
        elements.append(Paragraph("Reporte de Ventas", title_style))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Resumen Ejecutivo
        elements.append(Paragraph("Resumen Ejecutivo", heading_style))
        
        resumen = data['resumen_ejecutivo']
        resumen_data = [
            ['Métrica', 'Valor'],
            ['Total Ventas', f"${resumen['total_ventas']:,.0f} CLP"],
            ['Total Entradas', f"{resumen['total_entradas']:,}"],
            ['Promedio por Venta', f"${resumen['promedio_venta']:,.0f} CLP"],
        ]
        
        if resumen.get('sector_mas_vendido'):
            resumen_data.append(['Sector Más Vendido', resumen['sector_mas_vendido']])
        if resumen.get('sector_mayor_ingreso'):
            resumen_data.append(['Sector Mayor Ingreso', resumen['sector_mayor_ingreso']])
        
        resumen_table = Table(resumen_data, colWidths=[3*inch, 3*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(resumen_table)
        elements.append(Spacer(1, 20))
        
        # Análisis por Sector
        if data['analisis_por_sector']:
            elements.append(Paragraph("Análisis por Sector", heading_style))
            
            sector_data = [['Sector', 'Entradas', 'Ventas', 'Precio Prom.']]
            for sector, analisis in data['analisis_por_sector'].items():
                sector_data.append([
                    sector,
                    f"{analisis['entradas_vendidas']:,}",
                    f"${analisis['total_ventas']:,.0f}",
                    f"${analisis['precio_promedio']:,.0f}"
                ])
            
            sector_table = Table(sector_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            sector_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            elements.append(sector_table)
            elements.append(Spacer(1, 20))
        
        # Análisis por Evento
        if data['analisis_por_evento']:
            elements.append(Paragraph("Análisis por Evento", heading_style))
            
            evento_data = [['Evento', 'Entradas', 'Ventas Totales']]
            for evento, analisis in sorted(data['analisis_por_evento'].items(), 
                                          key=lambda x: x[1]['total_ventas'], 
                                          reverse=True):
                evento_data.append([
                    evento[:40] + '...' if len(evento) > 40 else evento,
                    f"{analisis['total_entradas']:,}",
                    f"${analisis['total_ventas']:,.0f}"
                ])
            
            evento_table = Table(evento_data, colWidths=[3*inch, 1.5*inch, 2*inch])
            evento_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (2, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            elements.append(evento_table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'reporte_ventas_{datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")}.pdf'
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error generando PDF: {str(e)}'}), 500


def generar_excel_reporte(data):
    """Genera un archivo Excel profesional con los datos del reporte"""
    try:
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=16, color="1e3a8a")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja 1: Resumen Ejecutivo
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"
        
        # Título
        ws1['A1'] = 'REPORTE DE VENTAS'
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:B1')
        
        ws1['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws1.merge_cells('A2:B2')
        
        # Resumen
        resumen = data['resumen_ejecutivo']
        row = 4
        ws1[f'A{row}'] = 'Métrica'
        ws1[f'B{row}'] = 'Valor'
        ws1[f'A{row}'].fill = header_fill
        ws1[f'B{row}'].fill = header_fill
        ws1[f'A{row}'].font = header_font
        ws1[f'B{row}'].font = header_font
        
        row += 1
        metricas = [
            ('Total Ventas', f"${resumen['total_ventas']:,.0f} CLP"),
            ('Total Entradas', f"{resumen['total_entradas']:,}"),
            ('Promedio por Venta', f"${resumen['promedio_venta']:,.0f} CLP"),
        ]
        
        if resumen.get('sector_mas_vendido'):
            metricas.append(('Sector Más Vendido', resumen['sector_mas_vendido']))
        if resumen.get('sector_mayor_ingreso'):
            metricas.append(('Sector Mayor Ingreso', resumen['sector_mayor_ingreso']))
        
        for metrica, valor in metricas:
            ws1[f'A{row}'] = metrica
            ws1[f'B{row}'] = valor
            ws1[f'A{row}'].border = border
            ws1[f'B{row}'].border = border
            row += 1
        
        # Ajustar ancho de columnas
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 30
        
        # Hoja 2: Análisis por Sector
        if data['analisis_por_sector']:
            ws2 = wb.create_sheet("Análisis por Sector")
            ws2['A1'] = 'ANÁLISIS POR SECTOR'
            ws2['A1'].font = title_font
            ws2.merge_cells('A1:D1')
            
            # Headers
            headers = ['Sector', 'Entradas Vendidas', 'Total Ventas', 'Precio Promedio']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=3, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            row = 4
            for sector, analisis in data['analisis_por_sector'].items():
                ws2[f'A{row}'] = sector
                ws2[f'B{row}'] = analisis['entradas_vendidas']
                ws2[f'C{row}'] = f"${analisis['total_ventas']:,.0f}"
                ws2[f'D{row}'] = f"${analisis['precio_promedio']:,.0f}"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws2[f'{col}{row}'].border = border
                row += 1
            
            # Ajustar anchos
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 18
            ws2.column_dimensions['C'].width = 18
            ws2.column_dimensions['D'].width = 18
        
        # Hoja 3: Datos Detallados
        if data['datos_detallados']:
            ws3 = wb.create_sheet("Datos Detallados")
            ws3['A1'] = 'VENTAS DETALLADAS'
            ws3['A1'].font = title_font
            ws3.merge_cells('A1:G1')
            
            # Headers
            headers = ['Fecha Venta', 'Evento', 'Cliente', 'Sector', 'Cantidad', 'Precio Unit.', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=3, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            row = 4
            for venta in data['datos_detallados'][:100]:  # Limitar a 100 registros
                ws3[f'A{row}'] = datetime.fromisoformat(venta['fecha_venta']).strftime('%d/%m/%Y %H:%M') if venta['fecha_venta'] else 'N/A'
                ws3[f'B{row}'] = venta['evento_nombre'][:30]
                ws3[f'C{row}'] = venta['cliente_nombre']
                ws3[f'D{row}'] = venta['sector_nombre']
                ws3[f'E{row}'] = venta['cantidad']
                ws3[f'F{row}'] = f"${venta['precio_unitario']:,.0f}"
                ws3[f'G{row}'] = f"${venta['total']:,.0f}"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws3[f'{col}{row}'].border = border
                row += 1
            
            # Ajustar anchos
            ws3.column_dimensions['A'].width = 18
            ws3.column_dimensions['B'].width = 25
            ws3.column_dimensions['C'].width = 20
            ws3.column_dimensions['D'].width = 15
            ws3.column_dimensions['E'].width = 10
            ws3.column_dimensions['F'].width = 15
            ws3.column_dimensions['G'].width = 15
        
        # Guardar y enviar
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'reporte_ventas_{datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")}.xlsx'
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error generando Excel: {str(e)}'}), 500

